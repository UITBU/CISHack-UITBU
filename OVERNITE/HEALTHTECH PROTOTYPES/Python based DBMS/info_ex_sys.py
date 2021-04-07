# -*- coding: utf-8 -*-
"""
Created on Wed Apr  7 20:14:18 2021

@author: Chirantan
"""

import openpyxl

loc = ("HEALTHDATA.xlsx")
wb_obj = openpyxl.load_workbook(loc)
k = wb_obj.active

f = open("demofile.txt", "rt")

class patient:
    flg = 0
    name = f.readline()
    ph = f.readline()
    bloodgr = f.readline()
    dob = f.readline()
    gender = f.readline()
    patid = "null";
    
    def update(X, id_no):
        if id_no != '' :
            X.patid = id_no
            X.flg = 1
        else:
            l1, ln, lp, ld = [], [], [], []
            for i in range(2, k.max_row + 1):
                l1.append(int(k.cell(row = i, column = 1).value))
                ln.append(int(k.cell(row = i, column = 2).value))
                lp.append(int(k.cell(row = i, column = 3).value))
                ld.append(int(k.cell(row = i, column = 5).value))
            
            j = str(max(l1) + 1)
            x = str(j).zfill(4)
            
            if X.name in ln:
                if lp[ln.index(X.name)] == X.ph or ld[ln.index(X.name)] == X.dob:
                    X.patid = l1[ln.index(X.name)]
                    X.flg = 1
                else:
                    X.patid = x
            else:
                X.patid = x
            
Y = patient()
Y.update(f.readline())

if Y.flg == 1:
    l1 = []
    for i in range(2, k.max_row + 1):
        l1.append(int(k.cell(row = i, column = 1).value))
    
    j = l1.index(Y.patid) + 1
    k.cell(row = j, column = 1).value = Y.patid
    k.cell(row = j, column = 2).value = Y.name
    k.cell(row = j, column = 3).value = Y.ph
    k.cell(row = j, column = 4).value = Y.bloodgr
    k.cell(row = i, column = 5).value = Y.dob
    k.cell(row = i, column = 6).value = Y.gender

else:
    k.cell(row = k.max_row + 1, column = 1).value = Y.patid
    k.cell(row = k.max_row + 1, column = 2).value = Y.name
    k.cell(row = k.max_row + 1, column = 3).value = Y.ph
    k.cell(row = k.max_row + 1, column = 4).value = Y.bloodgr
    k.cell(row = k.max_row + 1, column = 5).value = Y.dob
    k.cell(row = k.max_row + 1, column = 6).value = Y.gender
    
wb_obj.save("HEALTHDATA.xlsx");

f.close();
